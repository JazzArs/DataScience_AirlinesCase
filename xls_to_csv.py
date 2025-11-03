import glob
import os

import pandas as pd
from openpyxl import load_workbook

input_dir = "./unzipped_xlsx"
output_csv = "flights_parsed.csv"

COLUMNS = [
    "real_first_name", "real_last_name", "birth_date", "p_source",
    "flight_date", "flight_time", "flight_no", "codeshare",
    "dep_city", "dep_airport", "dep_country",
    "arr_city", "arr_airport", "arr_country",
    "e_code", "e_ticket", "docs", "seat", "meal",
    "booking_class", "fare_basis", "baggage", "loyalty_pairs"
]

rows = []

for path in glob.glob(os.path.join(input_dir, "*.xlsx")):
    wb = load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        full_name = str(ws.cell(row=3, column=2).value or "").strip()
        parts = full_name.split()

        if len(parts) >= 2:
            real_first_name = parts[0]
            real_last_name = parts[1]
        elif len(parts) == 1:
            real_first_name = parts[0]
            real_last_name = ""
        else:
            real_first_name = ""
            real_last_name = ""

        data = {
            "real_first_name": real_first_name,
            "real_last_name": real_last_name,
            "birth_date": "",
            "p_source": "xlsx",
            "flight_date": str(ws.cell(row=9, column=1).value or "").strip(),
            "flight_time": str(ws.cell(row=9, column=3).value or "").strip(),
            "flight_no": str(ws.cell(row=5, column=1).value or "").strip(),
            "codeshare": "",
            "dep_city": str(ws.cell(row=5, column=4).value or "").strip(),
            "dep_airport": str(ws.cell(row=7, column=4).value or "").strip(),
            "dep_country": "",
            "arr_city": str(ws.cell(row=5, column=8).value or "").strip(),
            "arr_airport": str(ws.cell(row=7, column=8).value or "").strip(),
            "arr_country": "",
            "e_code": str(ws.cell(row=13, column=2).value or "").strip(),
            "e_ticket": str(ws.cell(row=13, column=5).value or "").strip(),
            "docs": str(ws.cell(row=14, column=2).value or "").strip(),
            "seat": "N/A",
            "booking_class": str(ws.cell(row=3, column=8).value or "").strip(),
            "meal": "",
            "fare_basis": "",
            "baggage": "",
            "loyalty_pairs": ""
        }

        rows.append([data[col] for col in COLUMNS])

result_df = pd.DataFrame(rows, columns=COLUMNS)

result_df.to_csv(output_csv, sep=";", index=False, encoding="utf-8-sig")
